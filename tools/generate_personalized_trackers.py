"""
Generate personalized Excel trackers for Surendar and Arun
"""

import json
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation


class PersonalizedTrackerGenerator:
    def __init__(self, config_path, trader_name):
        self.config_path = config_path
        self.trader_name = trader_name
        with open(config_path) as f:
            self.config = json.load(f)
        self.wb = Workbook()
        self.wb.remove(self.wb.active)
        
    def create_tracker(self, output_path):
        """Create personalized Excel tracking workbook"""
        self.create_dashboard_sheet()
        self.create_daily_tracking_sheet()
        self.create_weekly_tracking_sheet()
        self.create_stats_sheet()
        self.wb.save(output_path)
        print(f"✅ {self.trader_name}'s tracker created: {output_path}")
        
    def get_color_for_category(self, category):
        """Get hex color for category"""
        colors = {
            "preparation": "4CAF50",
            "risk": "FF5252",
            "execution": "2196F3",
            "psychology": "FF9800",
            "review": "9C27B0",
            "analysis": "00BCD4",
            "education": "673AB7",
            "planning": "FF6F00",
            "wellness": "8BC34A"
        }
        return colors.get(category, "BDBDBD")
    
    def style_header(self, cell, color="2C3E50"):
        """Apply header styling"""
        cell.font = Font(bold=True, color="FFFFFF", size=12)
        cell.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    
    def style_data_cell(self, cell, bg_color="F5F5F5"):
        """Apply data cell styling"""
        cell.font = Font(size=11)
        cell.fill = PatternFill(start_color=bg_color, end_color=bg_color, fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
    
    def create_dashboard_sheet(self):
        """Create the main dashboard/summary sheet"""
        ws = self.wb.create_sheet("📊 Dashboard", 0)
        
        # Title with trader name
        ws.merge_cells('A1:H2')
        title = ws['A1']
        title.value = f"{self.trader_name}'S TRADING DISCIPLINE TRACKER"
        self.style_header(title, "1A4D2E")
        title.font = Font(bold=True, color="FFFFFF", size=16)
        
        # Subtitle with date
        ws.merge_cells('A3:H3')
        subtitle = ws['A3']
        subtitle.value = f"Week of {datetime.now().strftime('%B %d, %Y')}"
        subtitle.font = Font(italic=True, size=10)
        subtitle.alignment = Alignment(horizontal="center")
        
        # Summary boxes
        row = 5
        
        # Overall Progress
        ws.merge_cells(f'A{row}:B{row}')
        ws[f'A{row}'].value = "OVERALL PROGRESS"
        self.style_header(ws[f'A{row}'], "4CAF50")
        
        ws.merge_cells(f'A{row+1}:B{row+2}')
        progress_cell = ws[f'A{row+1}']
        progress_cell.value = "0%"
        progress_cell.font = Font(bold=True, size=24, color="4CAF50")
        progress_cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # Daily Habits Completed
        ws.merge_cells(f'D{row}:E{row}')
        ws[f'D{row}'].value = "DAILY HABITS"
        self.style_header(ws[f'D{row}'], "2196F3")
        
        ws.merge_cells(f'D{row+1}:E{row+2}')
        ws[f'D{row+1}'].value = "0/15"
        ws[f'D{row+1}'].font = Font(bold=True, size=24, color="2196F3")
        ws[f'D{row+1}'].alignment = Alignment(horizontal="center", vertical="center")
        
        # Weekly Habits Completed
        ws.merge_cells(f'G{row}:H{row}')
        ws[f'G{row}'].value = "WEEKLY HABITS"
        self.style_header(ws[f'G{row}'], "9C27B0")
        
        ws.merge_cells(f'G{row+1}:H{row+2}')
        ws[f'G{row+1}'].value = "0/8"
        ws[f'G{row+1}'].font = Font(bold=True, size=24, color="9C27B0")
        ws[f'G{row+1}'].alignment = Alignment(horizontal="center", vertical="center")
        
        # Category Summary
        row = 10
        ws.merge_cells(f'A{row}:H{row}')
        ws[f'A{row}'].value = "CATEGORY PROGRESS"
        self.style_header(ws[f'A{row}'], "37474F")
        
        row += 1
        categories = ["Preparation", "Risk Management", "Execution", "Psychology", "Review"]
        category_keys = ["preparation", "risk", "execution", "psychology", "review"]
        
        for col, (cat_name, cat_key) in enumerate(zip(categories, category_keys), 1):
            cell = ws.cell(row=row, column=col)
            cell.value = cat_name
            self.style_header(cell, self.get_color_for_category(cat_key))
            
        row += 1
        for col in range(1, 6):
            cell = ws.cell(row=row, column=col)
            cell.value = "0%"
            self.style_data_cell(cell)
            cell.font = Font(bold=True, size=12)
        
        # Set column widths
        for col in range(1, 9):
            ws.column_dimensions[get_column_letter(col)].width = 18
        
        ws.row_dimensions[1].height = 30
        ws.row_dimensions[3].height = 20
        
        return ws
    
    def create_daily_tracking_sheet(self):
        """Create daily habit tracking sheet with 7-day columns"""
        ws = self.wb.create_sheet("📋 Daily Tracking", 1)
        
        # Title
        ws.merge_cells('A1:H1')
        title = ws['A1']
        title.value = "DAILY HABIT TRACKING"
        self.style_header(title, "2196F3")
        title.font = Font(bold=True, color="FFFFFF", size=14)
        
        # Headers - Habit name + 7 days
        ws['A2'].value = "HABIT"
        ws['B2'].value = "CATEGORY"
        self.style_header(ws['A2'], "37474F")
        self.style_header(ws['B2'], "37474F")
        
        # Day headers (Mon-Sun)
        days = ["Mon", "Tue", "Wed", "Thu", "Fri", "Sat", "Sun"]
        for col, day in enumerate(days, 3):
            cell = ws.cell(row=2, column=col)
            cell.value = day
            self.style_header(cell, "1976D2")
        
        # Add habits
        row = 3
        daily_habits = self.config['daily_habits']
        
        for idx, habit in enumerate(daily_habits):
            # Habit name
            name_cell = ws[f'A{row}']
            name_cell.value = f"{habit['emoji']} {habit['name']}"
            name_cell.font = Font(size=11)
            name_cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
            self.style_data_cell(name_cell, "F5F5F5")
            
            # Category
            cat_cell = ws[f'B{row}']
            cat_cell.value = habit['category'].title()
            self.style_data_cell(cat_cell, self.get_color_for_category(habit['category']) + "40")
            
            # 7 day checkboxes
            for col in range(3, 10):
                cell = ws.cell(row=row, column=col)
                cell.value = ""
                self.style_data_cell(cell, "FFFFFF")
                dv = DataValidation(
                    type="list",
                    formula1='"✓,✗,"',
                    allow_blank=True
                )
                dv.add(cell)
            
            row += 1
        
        # Progress row
        ws[f'A{row}'].value = "DAILY PROGRESS"
        self.style_header(ws[f'A{row}'], "FF9800")
        ws[f'B{row}'].value = ""
        
        for col in range(3, 10):
            cell = ws.cell(row=row, column=col)
            cell.value = "0%"
            self.style_data_cell(cell, "FFE0B2")
            cell.font = Font(bold=True)
        
        # Set column widths
        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 15
        for col in range(3, 10):
            ws.column_dimensions[get_column_letter(col)].width = 12
        
        ws.row_dimensions[1].height = 25
        ws.row_dimensions[2].height = 20
        
        return ws
    
    def create_weekly_tracking_sheet(self):
        """Create weekly habit tracking sheet"""
        ws = self.wb.create_sheet("🎯 Weekly Tracking", 2)
        
        # Title
        ws.merge_cells('A1:C1')
        title = ws['A1']
        title.value = "WEEKLY HABIT TRACKING"
        self.style_header(title, "9C27B0")
        title.font = Font(bold=True, color="FFFFFF", size=14)
        
        # Headers
        headers = ["HABIT", "CATEGORY", "COMPLETED"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=2, column=col)
            cell.value = header
            self.style_header(cell, "37474F")
        
        # Add habits
        row = 3
        weekly_habits = self.config['weekly_habits']
        
        for habit in weekly_habits:
            # Habit name
            name_cell = ws[f'A{row}']
            name_cell.value = f"{habit['emoji']} {habit['name']}"
            name_cell.font = Font(size=11)
            self.style_data_cell(name_cell, "F5F5F5")
            
            # Category
            cat_cell = ws[f'B{row}']
            cat_cell.value = habit.get('category', 'Other').title()
            self.style_data_cell(cat_cell, "E1BEE7")
            
            # Completed checkbox
            check_cell = ws[f'C{row}']
            check_cell.value = ""
            self.style_data_cell(check_cell, "FFFFFF")
            dv = DataValidation(
                type="list",
                formula1='"✓,✗,"',
                allow_blank=True
            )
            dv.add(check_cell)
            
            row += 1
        
        # Summary
        row += 1
        ws[f'A{row}'].value = "WEEKLY PROGRESS"
        self.style_header(ws[f'A{row}'], "FF9800")
        ws[f'B{row}'].value = ""
        ws[f'C{row}'].value = "0/8"
        self.style_data_cell(ws[f'C{row}'], "FFE0B2")
        ws[f'C{row}'].font = Font(bold=True)
        
        # Set column widths
        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 20
        ws.column_dimensions['C'].width = 15
        
        ws.row_dimensions[1].height = 25
        ws.row_dimensions[2].height = 20
        
        return ws
    
    def create_stats_sheet(self):
        """Create statistics and analysis sheet"""
        ws = self.wb.create_sheet("📈 Statistics", 3)
        
        # Title
        ws.merge_cells('A1:D1')
        title = ws['A1']
        title.value = "TRACKING STATISTICS & INSIGHTS"
        self.style_header(title, "FF6F00")
        title.font = Font(bold=True, color="FFFFFF", size=14)
        
        # Key Metrics
        row = 3
        metrics = [
            ("Total Daily Habits", "15"),
            ("Total Weekly Habits", "8"),
            ("Current Week Progress", "0%"),
            ("Best Category", "—"),
            ("Needs Improvement", "—"),
            ("Current Streak", "0 days"),
            ("Overall Discipline Score", "0/100")
        ]
        
        for metric, value in metrics:
            label_cell = ws[f'A{row}']
            label_cell.value = metric
            label_cell.font = Font(bold=True, size=11)
            self.style_data_cell(label_cell, "F5F5F5")
            
            value_cell = ws[f'B{row}']
            value_cell.value = value
            value_cell.font = Font(size=11, color="1565C0")
            self.style_data_cell(value_cell, "E3F2FD")
            
            row += 1
        
        # Category Performance
        row += 2
        ws[f'A{row}'].value = "PERFORMANCE BY CATEGORY"
        self.style_header(ws[f'A{row}'], "37474F")
        
        row += 1
        categories = ["Preparation", "Risk Management", "Execution", "Psychology", "Review"]
        category_keys = ["preparation", "risk", "execution", "psychology", "review"]
        
        for cat_name, cat_key in zip(categories, category_keys):
            cell = ws[f'A{row}']
            cell.value = cat_name
            self.style_data_cell(cell, self.get_color_for_category(cat_key) + "40")
            
            cell = ws[f'B{row}']
            cell.value = "0%"
            self.style_data_cell(cell, "FFFFFF")
            
            row += 1
        
        # Instructions
        row += 2
        instr_title = ws[f'A{row}']
        instr_title.value = f"📝 {self.trader_name.upper()}'S TRACKER GUIDE:"
        instr_title.font = Font(bold=True, size=11, color="D32F2F")
        
        instructions = [
            "1. Go to the 'Daily Tracking' sheet each trading day",
            "2. Mark completed habits with ✓ (or ✗ if not completed)",
            "3. Go to 'Weekly Tracking' sheet and mark weekly tasks",
            "4. Progress automatically calculates in Dashboard",
            "5. Check 'Statistics' sheet for insights",
            "6. Aim for 85%+ consistent discipline score"
        ]
        
        for idx, instr in enumerate(instructions):
            row += 1
            cell = ws[f'A{row}']
            cell.value = instr
            cell.font = Font(size=10)
            cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        
        # Set column widths
        ws.column_dimensions['A'].width = 35
        ws.column_dimensions['B'].width = 25
        
        ws.row_dimensions[1].height = 25
        
        return ws


def generate_both_trackers():
    """Generate trackers for both Surendar and Arun"""
    config_path = "/Users/surendarnagarajan/Library/CloudStorage/OneDrive-Finastra/atraderview/Trading_Statergy/config/discipline_goals.json"
    base_dir = "/Users/surendarnagarajan/Library/CloudStorage/OneDrive-Finastra/atraderview/Trading_Statergy"
    
    # Create Surendar's tracker
    surendar_tracker = PersonalizedTrackerGenerator(config_path, "Surendar")
    surendar_tracker.create_tracker(f"{base_dir}/trackers/SURENDAR_DISCIPLINE_TRACKER.xlsx")
    
    # Create Arun's tracker
    arun_tracker = PersonalizedTrackerGenerator(config_path, "Arun")
    arun_tracker.create_tracker(f"{base_dir}/trackers/ARUN_DISCIPLINE_TRACKER.xlsx")
    
    print("\n✅ Both trackers created successfully!")
    print(f"   📊 Surendar's tracker: {base_dir}/trackers/SURENDAR_DISCIPLINE_TRACKER.xlsx")
    print(f"   📊 Arun's tracker: {base_dir}/trackers/ARUN_DISCIPLINE_TRACKER.xlsx")


if __name__ == "__main__":
    generate_both_trackers()
